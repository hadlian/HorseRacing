#!/usr/bin/env python3
"""
r5_analyze.py — R5 Performance Analysis + Excel Export

Usage:
    python3 r5_analyze.py                   # all results in DB
    python3 r5_analyze.py --track CD        # filter by track
    python3 r5_analyze.py --min-races 5     # need at least N races
    python3 r5_analyze.py --out report.xlsx # custom filename
"""

import argparse
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)

HORSE_RACING_ROOT = Path(__file__).resolve().parent.parent
DB_PATH           = HORSE_RACING_ROOT / "results" / "r5_results.db"
OUT_DIR           = HORSE_RACING_ROOT / "results"

# ── Colour palette ────────────────────────────────────────────────────────────
C_GOLD    = "C9A84C"
C_DARK    = "1A1A2E"
C_HEADER  = "16213E"
C_ALT     = "E8F4F8"
C_GREEN   = "27AE60"
C_RED     = "E74C3C"
C_ORANGE  = "E67E22"
C_WHITE   = "FFFFFF"
C_LIGHT   = "F5F5F5"

COMPONENTS = ["fci_n", "class_n", "bias_n", "tj_n", "form_n", "ped_n", "val_n",
              "best_dist_n", "pp_n"]
COMP_LABELS = {
    "fci_n":   "FCI (Speed/Trend)",
    "class_n": "Class vs Par",
    "bias_n":  "Bias/Pace Fit",
    "tj_n":    "Trainer/Jockey",
    "form_n":  "Form Angle",
    "ped_n":   "Pedigree",
    "val_n":   "Value vs ML",
    "best_dist_n": "Best @ Distance",
    "pp_n":    "Prime Power",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def get_conn():
    if not DB_PATH.exists():
        print(f"Database not found: {DB_PATH}")
        print("Run r5 with --track flag first to log picks.")
        sys.exit(1)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def hdr_fill(color):
    return PatternFill("solid", fgColor=color)


def hdr_font(color=C_WHITE, bold=True, size=11):
    return Font(color=color, bold=bold, size=size, name="Calibri")


def cell_font(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, name="Calibri", color=color)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def pct(num, den):
    return round(num / den * 100, 1) if den else 0.0


def roi(sp_payoffs, n_wins, total_bets):
    """ROI% for $2 flat win bets.
    sp_payoffs: list of $2 mutuel payoffs for winners with a recorded SP.
    Winners with no recorded payoff are counted as stake-back ($2 return)."""
    if not total_bets:
        return 0.0
    total_return = sum(sp_payoffs) + 2 * (n_wins - len(sp_payoffs))
    return round((total_return - 2 * total_bets) / (2 * total_bets) * 100, 1)


def safe_avg(values):
    v = [x for x in values if x is not None]
    return round(sum(v) / len(v), 2) if v else None


# ── Data loading ──────────────────────────────────────────────────────────────

def load_data(track_filter=None):
    conn = get_conn()
    where = "WHERE r.result_fetched=1"
    params = []
    if track_filter:
        where += " AND r.track=?"
        params.append(track_filter.upper())

    races = conn.execute(f"""
        SELECT r.id, r.track, r.date, r.race_num, r.surface, r.dist_f,
               r.race_type, r.purse, r.pace_scenario, r.speed_count
        FROM races r {where}
        ORDER BY r.date, CAST(r.race_num AS INT)
    """, params).fetchall()

    picks = {}
    for race in races:
        ps = conn.execute("""
            SELECT p.*
            FROM picks p
            WHERE p.race_id=?
            ORDER BY p.model_rank
        """, (race["id"],)).fetchall()
        picks[race["id"]] = [dict(p) for p in ps]

    conn.close()
    return [dict(r) for r in races], picks


# ── Analysis calculations ─────────────────────────────────────────────────────

def calc_summary(races, picks):
    n_races = len(races)
    # Exclude late scratches (finish_pos=-1) but keep NULL (ran, position not fully logged)
    all_picks = [p for ps in picks.values() for p in ps
                 if p.get("finish_pos") != -1]
    winners   = [p for p in all_picks if p.get("won") == 1]

    # Top pick (model_rank=1) performance — among horses that started (not late-scratched)
    top_picks = [p for p in all_picks if p.get("model_rank") == 1]
    top_wins  = [p for p in top_picks if p.get("won") == 1]
    top_win_sp = [p["sp_odds"] for p in top_wins if p.get("sp_odds")]

    # Top-3 model picks containing winner
    top3_races = 0
    for race_id, ps in picks.items():
        top3 = [p for p in ps if p.get("model_rank", 99) <= 3]
        if any(p.get("won") == 1 for p in top3):
            top3_races += 1

    # Value plays: val_n >= 7 and model_rank <= 5
    value_plays = [p for p in all_picks if (p.get("val_n") or 0) >= 7
                   and (p.get("model_rank", 99)) <= 5]
    value_wins  = [p for p in value_plays if p.get("won") == 1]
    value_sp    = [p["sp_odds"] for p in value_wins if p.get("sp_odds")]

    # HIGH tier top picks
    high_picks = [p for p in top_picks if p.get("tier") == "HIGH"]
    high_wins  = [p for p in high_picks if p.get("won") == 1]

    return {
        "n_races":        n_races,
        "n_horses":       len(all_picks),
        "top_pick_wins":  len(top_wins),
        "top_pick_pct":   pct(len(top_wins), len(top_picks)),
        "top3_hit":       top3_races,
        "top3_pct":       pct(top3_races, n_races),
        "value_plays":    len(value_plays),
        "value_wins":     len(value_wins),
        "value_pct":      pct(len(value_wins), len(value_plays)),
        "value_avg_sp":   safe_avg(value_sp),
        "value_roi":      roi(value_sp, len(value_wins), len(value_plays)),
        "high_picks":     len(high_picks),
        "high_wins":      len(high_wins),
        "high_pct":       pct(len(high_wins), len(high_picks)),
        "avg_sp_top":     safe_avg(top_win_sp),
    }


def calc_component_correlations(races, picks):
    """For each component, average score of winners vs non-winners."""
    rows = []
    all_picks = [p for ps in picks.values() for p in ps
                 if p.get("finish_pos") is not None and p["finish_pos"] > 0]
    for comp in COMPONENTS:
        winners = [p[comp] for p in all_picks if p.get("won") == 1 and p.get(comp) is not None]
        losers  = [p[comp] for p in all_picks if p.get("won") == 0 and p.get(comp) is not None]
        rows.append({
            "component": COMP_LABELS[comp],
            "avg_winner": safe_avg(winners),
            "avg_loser":  safe_avg(losers),
            "diff":       round((safe_avg(winners) or 0) - (safe_avg(losers) or 0), 2),
            "n_winners":  len(winners),
        })
    return sorted(rows, key=lambda r: r["diff"] or 0, reverse=True)


def calc_by_track(races, picks):
    from collections import defaultdict
    by_track = defaultdict(lambda: {"races": 0, "top_wins": 0, "top3": 0})
    for race in races:
        rid = race["id"]
        by_track[race["track"]]["races"] += 1
        ps = picks.get(rid, [])
        top = [p for p in ps if p.get("model_rank") == 1]
        if any(p.get("won") for p in top):
            by_track[race["track"]]["top_wins"] += 1
        top3 = [p for p in ps if p.get("model_rank", 99) <= 3]
        if any(p.get("won") for p in top3):
            by_track[race["track"]]["top3"] += 1

    rows = []
    for track, d in sorted(by_track.items()):
        rows.append({
            "track":    track,
            "races":    d["races"],
            "top_wins": d["top_wins"],
            "top_pct":  pct(d["top_wins"], d["races"]),
            "top3":     d["top3"],
            "top3_pct": pct(d["top3"], d["races"]),
        })
    return rows


def calc_pace_accuracy(races, picks):
    from collections import defaultdict
    by_scenario = defaultdict(lambda: {"races": 0, "speed_won": 0, "closer_won": 0, "top_win": 0})
    for race in races:
        rid  = race["id"]
        scen = race.get("pace_scenario") or "NORMAL"
        ps   = picks.get(rid, [])
        by_scenario[scen]["races"] += 1
        winner = next((p for p in ps if p.get("won") == 1), None)
        if winner:
            style = winner.get("pace_style", "")
            if style == "speed":
                by_scenario[scen]["speed_won"] += 1
            elif style == "closer":
                by_scenario[scen]["closer_won"] += 1
        top = next((p for p in ps if p.get("model_rank") == 1), None)
        if top and top.get("won"):
            by_scenario[scen]["top_win"] += 1

    rows = []
    for scen in ["HOT", "NORMAL", "SLOW"]:
        d = by_scenario.get(scen, {})
        n = d.get("races", 0)
        rows.append({
            "scenario":   scen,
            "races":      n,
            "speed_won":  d.get("speed_won", 0),
            "closer_won": d.get("closer_won", 0),
            "top_win":    d.get("top_win", 0),
            "top_pct":    pct(d.get("top_win", 0), n),
        })
    return rows


def calc_race_by_race(races, picks):
    rows = []
    for race in races:
        rid = race["id"]
        ps  = picks.get(rid, [])
        top = next((p for p in ps if p.get("model_rank") == 1), None)
        winner = next((p for p in ps if p.get("won") == 1), None)
        winner_rank = winner.get("model_rank") if winner else None
        rows.append({
            "date":       race["date"],
            "track":      race["track"],
            "race":       race["race_num"],
            "surface":    race["surface"],
            "dist_f":     race["dist_f"],
            "race_type":  race["race_type"],
            "scenario":   race.get("pace_scenario"),
            "top_horse":  top["horse_name"] if top else "—",
            "top_comp":   top["comp"] if top else None,
            "top_tier":   top["tier"] if top else None,
            "winner":     winner["horse_name"] if winner else "?",
            "winner_rank": winner_rank,
            "top_won":    1 if (top and top.get("won")) else 0,
            "top_sp":     winner.get("sp_odds") if winner else None,
        })
    return rows


def calc_value_roi(all_picks, min_val=6.0, step=0.5):
    """Return ROI by val_n threshold."""
    rows = []
    for threshold in [min_val + i * step for i in range(9)]:
        plays = [p for p in all_picks
                 if (p.get("val_n") or 0) >= threshold
                 and (p.get("model_rank", 99)) <= 5]
        wins  = [p for p in plays if p.get("won") == 1]
        sp_list = [p["sp_odds"] for p in wins if p.get("sp_odds")]
        avg_sp  = safe_avg(sp_list) or 0
        rows.append({
            "threshold": threshold,
            "plays":     len(plays),
            "wins":      len(wins),
            "win_pct":   pct(len(wins), len(plays)),
            "avg_sp":    avg_sp,
            "roi":       roi(sp_list, len(wins), len(plays)),
        })
    return rows


# ── Excel helpers ─────────────────────────────────────────────────────────────

def write_header_row(ws, row_num, values, bg_color=C_HEADER, fg_color=C_WHITE, height=22):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.fill = hdr_fill(bg_color)
        c.font = hdr_font(fg_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[row_num].height = height


def write_data_row(ws, row_num, values, alt=False, bold=False, fmt_map=None):
    fmt_map = fmt_map or {}
    fill = hdr_fill(C_ALT) if alt else hdr_fill(C_WHITE)
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.fill = fill
        c.font = cell_font(bold=bold)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        if col in fmt_map:
            c.number_format = fmt_map[col]


def colour_cell(c, val, good_threshold, bad_threshold=None, reverse=False):
    if val is None:
        return
    if not reverse:
        if val >= good_threshold:
            c.font = Font(bold=True, color=C_GREEN, name="Calibri", size=10)
        elif bad_threshold and val <= bad_threshold:
            c.font = Font(bold=True, color=C_RED, name="Calibri", size=10)
    else:
        if val <= good_threshold:
            c.font = Font(bold=True, color=C_GREEN, name="Calibri", size=10)
        elif bad_threshold and val >= bad_threshold:
            c.font = Font(bold=True, color=C_RED, name="Calibri", size=10)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws, title, subtitle=None, start_row=1, n_cols=10):
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=n_cols)
    c = ws.cell(row=start_row, column=1, value=title)
    c.fill = hdr_fill(C_DARK)
    c.font = Font(bold=True, size=16, color=C_GOLD, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 32

    if subtitle:
        ws.merge_cells(start_row=start_row+1, start_column=1,
                       end_row=start_row+1, end_column=n_cols)
        c2 = ws.cell(row=start_row+1, column=1, value=subtitle)
        c2.fill = hdr_fill(C_HEADER)
        c2.font = Font(size=10, color=C_GOLD, name="Calibri", italic=True)
        c2.alignment = Alignment(horizontal="center")
        ws.row_dimensions[start_row+1].height = 18
        return start_row + 2
    return start_row + 1


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_summary(wb, summary, by_track, pace_rows, generated_at):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    row = title_block(ws, "R5 HANDICAPPING — PERFORMANCE SUMMARY",
                      f"Generated {generated_at}", n_cols=6)
    row += 1

    # Key metrics grid
    metrics = [
        ("Races Analyzed",          summary["n_races"],         None,   None, ""),
        ("Top Pick Win Rate",        f"{summary['top_pick_pct']}%", 33.0, 20.0, "%"),
        ("Top-3 Contains Winner",    f"{summary['top3_pct']}%",  55.0, 35.0, "%"),
        ("Value Play Win Rate",      f"{summary['value_pct']}%", 25.0, 15.0, "%"),
        ("Value Play ROI",           f"{summary['value_roi']}%", 0,    None, "%"),
        ("HIGH Tier Win Rate",       f"{summary['high_pct']}%",  40.0, 25.0, "%"),
        ("Value Plays Found",        summary["value_plays"],     None,  None, ""),
        ("Avg SP (Top Pick Wins)",   summary["avg_sp_top"],      None,  None, "$"),
    ]

    write_header_row(ws, row, ["Metric", "Value", "Target", "Alert", "Note"])
    row += 1
    for i, (label, val, good, bad, note) in enumerate(metrics):
        target = f">{good}%" if good else "—"
        alert  = f"<{bad}%" if bad else "—"
        ws.cell(row=row, column=1, value=label).font = cell_font(bold=True)
        ws.cell(row=row, column=1).fill = hdr_fill(C_ALT if i % 2 else C_WHITE)
        c_val = ws.cell(row=row, column=2, value=val)
        c_val.fill = hdr_fill(C_ALT if i % 2 else C_WHITE)
        c_val.alignment = Alignment(horizontal="center")
        c_val.font = cell_font(bold=True)
        ws.cell(row=row, column=3, value=target).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=alert).alignment  = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=note).alignment   = Alignment(horizontal="center")
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border()
        row += 1

    row += 1

    # By-track table
    if by_track:
        write_header_row(ws, row, ["Track", "Races", "Top Pick W", "Top Pick %",
                                    "Top3 Hit", "Top3 %"], bg_color=C_GOLD, fg_color=C_DARK)
        row += 1
        for i, t in enumerate(by_track):
            vals = [t["track"], t["races"], t["top_wins"],
                    f"{t['top_pct']}%", t["top3"], f"{t['top3_pct']}%"]
            write_data_row(ws, row, vals, alt=(i % 2 == 1))
            colour_cell(ws.cell(row=row, column=4), t["top_pct"], 33, 20)
            row += 1

    row += 1

    # Pace accuracy table
    if pace_rows:
        write_header_row(ws, row, ["Scenario", "Races", "Speed Won",
                                    "Closer Won", "Top Pick Wins", "Top Pick %"],
                         bg_color=C_GOLD, fg_color=C_DARK)
        row += 1
        for i, p in enumerate(pace_rows):
            vals = [p["scenario"], p["races"], p["speed_won"],
                    p["closer_won"], p["top_win"], f"{p['top_pct']}%"]
            write_data_row(ws, row, vals, alt=(i % 2 == 1))
            row += 1

    set_col_widths(ws, [26, 12, 14, 14, 12, 12])
    ws.freeze_panes = "A4"


def build_race_by_race(wb, rbr_rows):
    ws = wb.create_sheet("Race by Race")
    ws.sheet_view.showGridLines = False

    row = title_block(ws, "RACE-BY-RACE RESULTS", n_cols=14)
    row += 1

    headers = ["Date", "Track", "Race", "Surface", "Dist(f)", "Type",
               "Scenario", "Top Pick", "Score", "Tier",
               "Winner", "Winner Rank", "✓", "SP"]
    write_header_row(ws, row, headers)
    row += 1

    for i, r in enumerate(rbr_rows):
        won = r["top_won"]
        vals = [
            r["date"], r["track"], r["race"], r["surface"],
            r["dist_f"], r["race_type"], r["scenario"],
            r["top_horse"], r["top_comp"], r["top_tier"],
            r["winner"], r["winner_rank"], "✓" if won else "✗",
            r["top_sp"],
        ]
        write_data_row(ws, row, vals, alt=(i % 2 == 1),
                       fmt_map={9: "0.00", 14: "0.00"})

        # Colour the tick
        c = ws.cell(row=row, column=13)
        c.font = Font(bold=True,
                      color=C_GREEN if won else C_RED,
                      name="Calibri", size=11)
        c.alignment = Alignment(horizontal="center")

        # Colour tier
        tier_cell = ws.cell(row=row, column=10)
        tier_colors = {"HIGH": C_GREEN, "SOLID": "2980B9",
                       "FAIR": C_ORANGE, "SPECULATIVE": C_RED}
        if r["top_tier"] in tier_colors:
            tier_cell.font = Font(bold=True,
                                  color=tier_colors[r["top_tier"]],
                                  name="Calibri", size=10)
        row += 1

    set_col_widths(ws, [11, 7, 6, 9, 7, 10, 9, 22, 7, 10, 22, 12, 5, 7])
    ws.freeze_panes = "A4"


def build_component_correlations(wb, corr_rows):
    ws = wb.create_sheet("Component Correlations")
    ws.sheet_view.showGridLines = False

    row = title_block(ws, "COMPONENT CORRELATION WITH WINNERS",
                      "Higher diff = component better distinguishes winners from losers", n_cols=5)
    row += 1

    headers = ["Component", "Avg Score (Winner)", "Avg Score (Loser)", "Difference", "Winner Count"]
    write_header_row(ws, row, headers)
    row += 1

    for i, c in enumerate(corr_rows):
        vals = [c["component"], c["avg_winner"], c["avg_loser"],
                c["diff"], c["n_winners"]]
        write_data_row(ws, row, vals, alt=(i % 2 == 1),
                       fmt_map={2: "0.00", 3: "0.00", 4: "0.00"})
        diff_cell = ws.cell(row=row, column=4)
        colour_cell(diff_cell, c["diff"], 0.3, 0.0)
        row += 1

    # Bar chart
    if len(corr_rows) >= 2:
        chart = BarChart()
        chart.type  = "col"
        chart.title = "Component Differentiation (Winner vs Loser)"
        chart.y_axis.title = "Avg Score"
        chart.x_axis.title = "Component"
        chart.height = 14
        chart.width  = 22

        data_start = row - len(corr_rows)
        cats  = Reference(ws, min_col=1, min_row=data_start,
                          max_row=data_start + len(corr_rows) - 1)
        w_ref = Reference(ws, min_col=2, min_row=data_start - 1,
                          max_row=data_start + len(corr_rows) - 1)
        l_ref = Reference(ws, min_col=3, min_row=data_start - 1,
                          max_row=data_start + len(corr_rows) - 1)
        chart.add_data(w_ref, titles_from_data=True)
        chart.add_data(l_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{row + 2}")

    set_col_widths(ws, [24, 18, 18, 14, 14])


def build_value_roi(wb, value_rows):
    ws = wb.create_sheet("Value ROI")
    ws.sheet_view.showGridLines = False

    row = title_block(ws, "VALUE PLAY ROI BY THRESHOLD",
                      "Value plays = val_n >= threshold AND model_rank <= 5", n_cols=6)
    row += 1

    headers = ["Val_n Threshold", "Total Plays", "Wins", "Win %", "Avg SP", "ROI %"]
    write_header_row(ws, row, headers)
    row += 1

    chart_start = row
    for i, v in enumerate(value_rows):
        vals = [v["threshold"], v["plays"], v["wins"],
                v["win_pct"], v["avg_sp"], v["roi"]]
        write_data_row(ws, row, vals, alt=(i % 2 == 1),
                       fmt_map={5: "0.00", 6: "0.0"})

        roi_cell = ws.cell(row=row, column=6)
        colour_cell(roi_cell, v["roi"], 0, -20)
        win_pct_cell = ws.cell(row=row, column=4)
        colour_cell(win_pct_cell, v["win_pct"], 25, 10)
        row += 1

    # Line chart for ROI
    if len(value_rows) >= 3:
        chart = LineChart()
        chart.title  = "ROI % by Value Threshold"
        chart.y_axis.title = "ROI %"
        chart.x_axis.title = "Val_n Threshold"
        chart.height = 12
        chart.width  = 18

        roi_ref  = Reference(ws, min_col=6, min_row=chart_start - 1,
                             max_row=chart_start + len(value_rows) - 1)
        cats_ref = Reference(ws, min_col=1, min_row=chart_start,
                             max_row=chart_start + len(value_rows) - 1)
        chart.add_data(roi_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{row + 2}")

    set_col_widths(ws, [16, 12, 8, 10, 10, 10])


def build_scout_impact(wb, picks):
    ws = wb.create_sheet("Scout Impact")
    ws.sheet_view.showGridLines = False

    all_picks = [p for ps in picks.values() for p in ps
                 if p.get("finish_pos") is not None and p["finish_pos"] > 0]

    row = title_block(ws, "SCOUT ADJUSTMENT IMPACT",
                      "Horses with vs without scout adjustments, among winners", n_cols=5)
    row += 1

    adjusted = [p for p in all_picks if (p.get("scout_adj") or 0) != 0]
    non_adj   = [p for p in all_picks if (p.get("scout_adj") or 0) == 0]

    adj_wins = [p for p in adjusted if p.get("won") == 1]
    non_wins = [p for p in non_adj   if p.get("won") == 1]

    headers = ["Group", "Horses", "Winners", "Win Rate"]
    write_header_row(ws, row, headers)
    row += 1

    for i, (label, total, wins) in enumerate([
        ("With Scout Adj",    len(adjusted), len(adj_wins)),
        ("No Scout Adj",      len(non_adj),  len(non_wins)),
        ("Positive Adj Only", len([p for p in adjusted if (p.get("scout_adj") or 0) > 0]),
                              len([p for p in adjusted if (p.get("scout_adj") or 0) > 0 and p.get("won")])),
        ("Negative Adj Only", len([p for p in adjusted if (p.get("scout_adj") or 0) < 0]),
                              len([p for p in adjusted if (p.get("scout_adj") or 0) < 0 and p.get("won")])),
    ]):
        vals = [label, total, wins, f"{pct(wins, total)}%"]
        write_data_row(ws, row, vals, alt=(i % 2 == 1))
        colour_cell(ws.cell(row=row, column=4), pct(wins, total), 25, 10)
        row += 1

    row += 2
    # Detailed breakdown by adjustment type
    write_header_row(ws, row, ["Horse", "Race", "Scout Adj", "Comp Score", "Winner"])
    row += 1
    for i, p in enumerate(sorted(adjusted, key=lambda x: abs(x.get("scout_adj") or 0), reverse=True)[:30]):
        race_label = f"R{p.get('race_id', '?')}"
        vals = [p.get("horse_name"), race_label, p.get("scout_adj"),
                p.get("comp"), "✓" if p.get("won") else "✗"]
        write_data_row(ws, row, vals, alt=(i % 2 == 1), fmt_map={3: "+0.00", 4: "0.00"})
        tick = ws.cell(row=row, column=5)
        tick.font = Font(bold=True,
                         color=C_GREEN if p.get("won") else C_RED,
                         name="Calibri", size=10)
        row += 1

    set_col_widths(ws, [24, 10, 12, 12, 10])


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="R5 Performance Analysis")
    ap.add_argument("--track",     help="Filter by track code (e.g. CD)")
    ap.add_argument("--min-races", type=int, default=1,
                    help="Minimum completed races required (default: 1)")
    ap.add_argument("--out",       help="Output filename (default: auto-generated)")
    args = ap.parse_args()

    races, picks = load_data(track_filter=args.track)

    if len(races) < args.min_races:
        print(f"Only {len(races)} completed races found "
              f"(need {args.min_races}). Use r5_tracker.py to log results.")
        sys.exit(0)

    print(f"\n📊 Analyzing {len(races)} races...")

    summary  = calc_summary(races, picks)
    by_track = calc_by_track(races, picks)
    corr     = calc_component_correlations(races, picks)
    pace     = calc_pace_accuracy(races, picks)
    rbr      = calc_race_by_race(races, picks)
    all_picks = [p for ps in picks.values() for p in ps
                 if p.get("finish_pos") is not None and p["finish_pos"] > 0]
    val_roi  = calc_value_roi(all_picks)

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Print quick summary to console
    print(f"\n{'='*52}")
    print(f"  R5 ANALYSIS  ({summary['n_races']} races, {summary['n_horses']} horses)")
    print(f"{'='*52}")
    print(f"  Top pick win rate:    {summary['top_pick_pct']}%")
    print(f"  Top-3 hit rate:       {summary['top3_pct']}%")
    print(f"  Value play ROI:       {summary['value_roi']}%")
    print(f"  HIGH tier win rate:   {summary['high_pct']}%")
    print()
    print("  Component differentiation (winner avg vs loser avg):")
    for c in corr:
        bar = "+" * min(20, max(0, int(c["diff"] * 10)))
        print(f"    {c['component']:<22} diff={c['diff']:+.2f}  {bar}")
    print()

    # Build Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    build_summary(wb, summary, by_track, pace, generated_at)
    build_race_by_race(wb, rbr)
    build_component_correlations(wb, corr)
    build_value_roi(wb, val_roi)
    build_scout_impact(wb, picks)

    # Save
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if args.out:
        out_path = Path(args.out)
    else:
        suffix = f"_{args.track.upper()}" if args.track else ""
        out_path = OUT_DIR / f"r5_analysis{suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    wb.save(out_path)
    print(f"  💾 Saved → {out_path}")


if __name__ == "__main__":
    main()
