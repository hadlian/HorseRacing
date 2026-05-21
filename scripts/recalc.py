"""
recalc.py — Excel formula validator.
Scans an xlsx file for error values and malformed formula strings.
Returns JSON: {"status": "success"|"error", "errors": [...], "formula_count": N}
"""

import json
import sys
import openpyxl

ERROR_VALUES = {'#VALUE!', '#REF!', '#NAME?', '#DIV/0!', '#N/A', '#NULL!', '#NUM!', '#ERROR!'}


def validate(xlsx_path: str) -> dict:
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        return {"status": "error", "errors": [f"Cannot open file: {e}"], "formula_count": 0}

    errors = []
    formula_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                s = str(val)
                # Check for error values stored as strings
                if s.strip().upper() in ERROR_VALUES:
                    errors.append(f"{sheet_name}!{cell.coordinate}: {val}")
                # Count formulas and do basic bracket check
                if s.startswith('='):
                    formula_count += 1
                    if s.count('(') != s.count(')'):
                        errors.append(f"{sheet_name}!{cell.coordinate}: unmatched parentheses in formula: {s[:80]}")

    result = {
        "status": "success" if not errors else "error",
        "formula_count": formula_count,
        "errors": errors,
    }
    return result


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "errors": ["No file specified"], "formula_count": 0}))
        sys.exit(1)

    result = validate(sys.argv[1])
    print(json.dumps(result, indent=2))
    if result["status"] != "success":
        sys.exit(1)
