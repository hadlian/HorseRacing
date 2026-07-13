"""
CompareModels head-to-head comparison report generator.
Produces a 5-sheet Excel workbook comparing CM vs R5 on the 63-race universe.
"""

import json
import os
import sqlite3
import subprocess
import sys
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'Claude'))
from r5_paths import R5_DB_PATH  # noqa: E402

CM_DB = os.path.join(os.path.dirname(__file__), 'comparemodels_results.db')
R5_DB = str(R5_DB_PATH)
REPORTS_DIR = os.path.join(os.path.dirname(__file__), 'reports')
RECALC = os.path.join(os.path.dirname(__file__), '..', 'scripts', 'recalc.py')


# --- Styles ---
HDR_FILL_BLUE  = PatternFill("solid", fgColor="2F5496")
HDR_FILL_GREEN = PatternFill("solid", fgColor="375623")
HDR_FILL_GREY  = PatternFill("solid", fgColor="595959")
HDR_FILL_PURP  = PatternFill("solid", fgColor="7030A0")
HDR_FILL_DARK  = PatternFill("solid", fgColor="1F3864")
HDR_FONT_WHITE = Font(bold=True, color="FFFFFF")
BOLD           = Font(bold=True)
THIN           = Side(border_style="thin", color="CCCCCC")
THIN_BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr_cell(ws, row, col, value, fill=None, font=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    c.font = font or HDR_FONT_WHITE
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    return c


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --- Data loading ---

def load_sheet2_data() -> list[dict]:
    """
    Load all 63 aligned race rows from both DBs.
    One row per race (aligned on track+date+race_num).
    """
    r5_con = sqlite3.connect(f"file:{os.path.abspath(R5_DB)}?mode=ro", uri=True)
    r5_con.row_factory = sqlite3.Row
    cm_con = sqlite3.connect(f"file:{os.path.abspath(CM_DB)}?mode=ro", uri=True)
    cm_con.row_factory = sqlite3.Row

    r5_cur = r5_con.cursor()
    cm_cur = cm_con.cursor()

    # Load all R5 races (result_fetched=1)
    r5_cur.execute("""
        SELECT r.track, r.date, CAST(r.race_num AS INTEGER) as race_num,
               r.surface, r.dist_f, r.race_type, r.pace_scenario,
               p.pgm as r5_pgm, p.model_rank as r5_rank, p.comp as r5_score,
               p.tier as r5_tier, p.finish_pos as r5_finish,
               p.sp_odds as r5_sp, p.won as r5_won, p.ml_odds as r5_ml,
               p.horse_name as r5_winner_name
        FROM races r
        JOIN picks p ON p.race_id = r.id AND p.model_rank = 1
        WHERE r.result_fetched = 1
        ORDER BY r.date, r.track, CAST(r.race_num AS INTEGER)
    """)
    r5_top = {(row['track'], row['date'], row['race_num']): dict(row)
              for row in r5_cur.fetchall()}

    # Load winner names for R5
    r5_cur.execute("""
        SELECT r.track, r.date, CAST(r.race_num AS INTEGER) as rn,
               p.horse_name, p.finish_pos, p.model_rank, p.pgm
        FROM picks p JOIN races r ON p.race_id = r.id
        WHERE r.result_fetched = 1 AND p.won = 1
    """)
    r5_winners = {(row[0], row[1], row[2]): {'name': row[3], 'r5_rank': row[5], 'pgm': row[6]}
                  for row in r5_cur.fetchall()}

    # CM top picks
    cm_cur.execute("""
        SELECT track, race_date, race,
               horse_pgm as cm_pgm, horse_name as cm_name,
               cm_rank, composite_score as cm_score, tier as cm_tier,
               consensus_count, morning_line as cm_ml
        FROM picks
        WHERE cm_rank = 1
        ORDER BY race_date, track, race
    """)
    cm_top = {(row['track'], row['race_date'], row['race']): dict(row)
              for row in cm_cur.fetchall()}

    # CM results for top picks
    cm_cur.execute("""
        SELECT track, race_date, race, horse_pgm,
               finish_position, sp_odds
        FROM results
    """)
    cm_results_all = {}
    for row in cm_cur.fetchall():
        cm_results_all[(row[0], row[1], row[2], row[3])] = {'finish': row[4], 'sp': row[5]}

    # CM winner rank (for the race winner — find who won and their CM rank)
    cm_cur.execute("""
        SELECT p.track, p.race_date, p.race, p.horse_pgm, p.cm_rank
        FROM picks p
        JOIN results r ON p.track=r.track AND p.race_date=r.race_date
                       AND p.race=r.race AND p.horse_pgm=r.horse_pgm
        WHERE r.finish_position = 1
    """)
    cm_winner_ranks = {(row[0], row[1], row[2]): row[4] for row in cm_cur.fetchall()}

    # Load actual race winner SP from R5 (won=1 row) — needed for CM SP ROI
    r5_cur.execute("""
        SELECT r.track, r.date, CAST(r.race_num AS INTEGER), p.sp_odds
        FROM picks p JOIN races r ON p.race_id=r.id
        WHERE p.won=1 AND r.result_fetched=1
    """)
    race_winner_sp = {(row[0], row[1], row[2]): row[3] for row in r5_cur.fetchall()}

    rows = []
    for key, r5_row in sorted(r5_top.items(), key=lambda x: (x[0][1], x[0][0], x[0][2])):
        track, date, race_num = key
        cm = cm_top.get(key, {})

        winner = r5_winners.get(key, {})
        winner_name = winner.get('name', '')
        winner_r5_rank = winner.get('r5_rank', '')
        winner_cm_rank = cm_winner_ranks.get(key, '')

        cm_pgm = cm.get('cm_pgm', '')
        r5_pgm = r5_row.get('r5_pgm', '')

        agreement = 'Y' if cm_pgm and r5_pgm and cm_pgm == r5_pgm else 'N'

        # SP: actual race winner's post-time odds (from R5 won=1 row)
        winner_sp = race_winner_sp.get(key)
        # r5_sp: SP on R5 model_rank=1 row (only populated when R5 top pick won)
        r5_sp_val = r5_row.get('r5_sp')

        # R5 ROI: top pick won?
        r5_won_flag = 1 if r5_row.get('r5_won') == 1 else 0
        r5_ml = r5_row.get('r5_ml')
        # ML profit on $2 win bet at X-1 odds = 2X; SP column stores $2 mutuel
        # payoff, so profit = payoff - 2. (Corrected 2026-06-11 — old formulas
        # treated the payoff as decimal odds, inflating SP ROI ~2x.)
        r5_roi    = (r5_ml * 2) if r5_won_flag and r5_ml else -2
        r5_sp_roi = (r5_sp_val - 2) if r5_won_flag and r5_sp_val else -2

        # CM ROI: did CM top pick win?
        cm_result = cm_results_all.get((track, date, race_num, cm_pgm), {})
        cm_finish = cm_result.get('finish')
        cm_won = 1 if cm_finish == 1 else 0
        cm_ml = cm.get('cm_ml')
        cm_roi    = (cm_ml * 2) if cm_won and cm_ml else -2
        # CM SP ROI uses actual winner SP (same horse pays same regardless of who picked it)
        cm_sp_roi = (winner_sp - 2) if cm_won and winner_sp else -2

        rows.append({
            'date':         date,
            'track':        track,
            'race':         race_num,
            'surface':      r5_row.get('surface', ''),
            'dist_f':       r5_row.get('dist_f', ''),
            'race_type':    r5_row.get('race_type', ''),
            'pace_scenario':r5_row.get('pace_scenario', ''),
            'r5_pgm':       r5_pgm,
            'r5_tier':      r5_row.get('r5_tier', ''),
            'r5_score':     r5_row.get('r5_score', ''),
            'cm_pgm':       cm_pgm,
            'cm_tier':      cm.get('cm_tier', ''),
            'cm_score':     cm.get('cm_score', ''),
            'cm_consensus': cm.get('consensus_count', ''),
            'agreement':    agreement,
            'winner_name':  winner_name,
            'winner_r5_rank': winner_r5_rank,
            'winner_cm_rank': winner_cm_rank,
            'sp':           winner_sp,
            'r5_won':       r5_won_flag,
            'cm_won':       cm_won,
            'r5_ml':        r5_ml,
            'cm_ml':        cm_ml,
            'r5_roi':       r5_roi,
            'cm_roi':       cm_roi,
            'r5_sp_roi':    r5_sp_roi,
            'cm_sp_roi':    cm_sp_roi,
        })

    r5_con.close()
    cm_con.close()
    return rows


def load_cm_signals() -> dict:
    """Load signal stats for Sheet 4."""
    cm = sqlite3.connect(f"file:{os.path.abspath(CM_DB)}?mode=ro", uri=True)
    cm.row_factory = sqlite3.Row
    cur = cm.cursor()

    signals = {}

    # A-tier win rate
    cur.execute("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN r.finish_position=1 THEN 1 ELSE 0 END) as wins
        FROM picks p
        JOIN results r ON p.track=r.track AND p.race_date=r.race_date
                       AND p.race=r.race AND p.horse_pgm=r.horse_pgm
        WHERE p.tier='A'
    """)
    row = cur.fetchone()
    signals['a_tier'] = {'total': row[0], 'wins': row[1],
                         'rate': row[1]/row[0]*100 if row[0] else 0}

    # Dominant flag win rate
    cur.execute("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN r.finish_position=1 THEN 1 ELSE 0 END) as wins
        FROM picks p
        JOIN results r ON p.track=r.track AND p.race_date=r.race_date
                       AND p.race=r.race AND p.horse_pgm=r.horse_pgm
        WHERE p.is_dominant=1
    """)
    row = cur.fetchone()
    signals['dominant'] = {'total': row[0], 'wins': row[1],
                           'rate': row[1]/row[0]*100 if row[0] else 0}

    # Overlay Watch win rate
    cur.execute("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN r.finish_position=1 THEN 1 ELSE 0 END) as wins,
               SUM(CASE WHEN r.finish_position=1 THEN (p.morning_line*2) ELSE -2 END) as profit,
               COUNT(*)*2 as invested
        FROM picks p
        JOIN results r ON p.track=r.track AND p.race_date=r.race_date
                       AND p.race=r.race AND p.horse_pgm=r.horse_pgm
        WHERE p.is_overlay=1
    """)
    row = cur.fetchone()
    ov_roi = (row[2] / row[3] * 100) if row[3] and row[3] > 0 else 0
    signals['overlay'] = {'total': row[0], 'wins': row[1],
                          'rate': row[1]/row[0]*100 if row[0] else 0,
                          'roi': ov_roi}

    # Consensus count vs win rate (levels 1-8) — for CM rank=1 horses
    cons_stats = {}
    for level in range(1, 9):
        cur.execute("""
            SELECT COUNT(*) as total,
                   SUM(CASE WHEN r.finish_position=1 THEN 1 ELSE 0 END) as wins
            FROM picks p
            JOIN results r ON p.track=r.track AND p.race_date=r.race_date
                           AND p.race=r.race AND p.horse_pgm=r.horse_pgm
            WHERE p.consensus_count=? AND p.cm_rank=1
        """, (level,))
        row = cur.fetchone()
        cons_stats[level] = {'total': row[0], 'wins': row[1],
                             'rate': row[1]/row[0]*100 if row[0] else 0}
    signals['consensus'] = cons_stats

    # Per-category underline hit rate
    cat_stats = {}
    cur.execute("""
        SELECT cp.category,
               COUNT(*) as total,
               SUM(CASE WHEN r.finish_position=1 THEN 1 ELSE 0 END) as wins
        FROM category_picks cp
        JOIN results r ON cp.track=r.track AND cp.race_date=r.race_date
                       AND cp.race=r.race AND cp.horse_pgm=r.horse_pgm
        WHERE cp.underlined=1 AND cp.rank_in_cat=1
        GROUP BY cp.category
    """)
    for row in cur.fetchall():
        cat_stats[row[0]] = {'total': row[1], 'wins': row[2],
                             'rate': row[2]/row[1]*100 if row[1] else 0}
    signals['underline'] = cat_stats

    cm.close()
    return signals


def load_disagreements(rows: list[dict]) -> list[dict]:
    return [r for r in rows if r['agreement'] == 'N']


def build_report(output_path: str) -> str:
    """Build the 5-sheet comparison workbook. Returns output path."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    data = load_sheet2_data()
    signals = load_cm_signals()
    n = len(data)

    wb = openpyxl.Workbook()

    # ================================================================
    # SHEET 2 — Race by Race (build first; Sheet 1 references it)
    # ================================================================
    ws2 = wb.active
    ws2.title = "Race by Race"

    s2_headers = [
        'Date', 'Track', 'Race', 'Surface', 'Distance', 'Race Type',
        'Pace Scenario', 'R5 Pick', 'R5 Tier', 'R5 Score',
        'CM Pick', 'CM Tier', 'CM Composite', 'CM Consensus',
        'Agreement', 'Winner Name', 'Winner R5 Rank', 'Winner CM Rank',
        'SP', 'R5 Win', 'CM Win', 'R5 ML', 'CM ML', 'R5 ROI', 'CM ROI',
        'R5 SP ROI', 'CM SP ROI',
    ]
    for ci, h in enumerate(s2_headers, 1):
        hdr_cell(ws2, 1, ci, h, HDR_FILL_DARK)

    for ri, row in enumerate(data, 2):
        ws2.cell(ri, 1, row['date'])
        ws2.cell(ri, 2, row['track'])
        ws2.cell(ri, 3, row['race'])
        ws2.cell(ri, 4, row['surface'])
        ws2.cell(ri, 5, row['dist_f'])
        ws2.cell(ri, 6, row['race_type'])
        ws2.cell(ri, 7, row['pace_scenario'])
        ws2.cell(ri, 8, row['r5_pgm'])
        ws2.cell(ri, 9, row['r5_tier'])
        ws2.cell(ri, 10, row['r5_score'])
        ws2.cell(ri, 11, row['cm_pgm'])
        ws2.cell(ri, 12, row['cm_tier'])
        ws2.cell(ri, 13, row['cm_score'])
        ws2.cell(ri, 14, row['cm_consensus'])
        ws2.cell(ri, 15, row['agreement'])
        ws2.cell(ri, 16, row['winner_name'])
        ws2.cell(ri, 17, row['winner_r5_rank'])
        ws2.cell(ri, 18, row['winner_cm_rank'])
        ws2.cell(ri, 19, row['sp'])
        ws2.cell(ri, 20, row['r5_won'])
        ws2.cell(ri, 21, row['cm_won'])
        ws2.cell(ri, 22, row['r5_ml'])
        ws2.cell(ri, 23, row['cm_ml'])
        ws2.cell(ri, 24, row['r5_roi'])
        ws2.cell(ri, 25, row['cm_roi'])
        ws2.cell(ri, 26, row['r5_sp_roi'])
        ws2.cell(ri, 27, row['cm_sp_roi'])

    set_col_widths(ws2, [12, 6, 5, 7, 7, 14, 12,
                          8, 8, 8, 8, 8, 10, 10,
                          9, 22, 12, 12,
                          8, 7, 7, 8, 8, 8, 8, 9, 9])
    ws2.freeze_panes = 'A2'

    last_row = n + 1  # last data row (1-indexed, row 2 = first data)

    # Column letters in Sheet 2 (1-indexed matches above)
    # R5 Win = col T (20), CM Win = col U (21)
    # R5 ROI col = X (24), CM ROI = Y (25), R5 SP ROI = Z (26), CM SP ROI = AA (27)
    # R5 ML = V (22), CM ML = W (23)
    # Agreement = O (15), R5 Tier = I (9), CM Tier = L (12)
    # SP = S (19)

    def s2_range(col_letter):
        return f"'Race by Race'!{col_letter}2:{col_letter}{last_row}"

    # ================================================================
    # SHEET 1 — Summary (Excel formulas referencing Sheet 2)
    # ================================================================
    ws1 = wb.create_sheet("Summary", 0)
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18

    hdr_cell(ws1, 1, 1, 'Metric',        HDR_FILL_BLUE)
    hdr_cell(ws1, 1, 2, 'CompareModels', HDR_FILL_GREEN)
    hdr_cell(ws1, 1, 3, 'R5',            HDR_FILL_PURP)

    metrics = [
        ('Top Pick Win Rate (%)',
         f"=COUNTIF({s2_range('U')},1)/COUNTA({s2_range('K')})*100",
         f"=COUNTIF({s2_range('T')},1)/COUNTA({s2_range('H')})*100"),
        ('Top-3 Hit Rate (%)',
         f"=COUNTIF({s2_range('R')},\"<=3\")/COUNTA({s2_range('K')})*100",
         f"=COUNTIF({s2_range('Q')},\"<=3\")/COUNTA({s2_range('H')})*100"),
        ('Avg SP on Top Pick Wins',
         f"=AVERAGEIF({s2_range('U')},1,{s2_range('S')})",
         f"=AVERAGEIF({s2_range('T')},1,{s2_range('S')})"),
        ('ROI on ML — $2 win (%)',
         f"=SUM({s2_range('Y')})/({n}*2)*100",
         f"=SUM({s2_range('X')})/({n}*2)*100"),
        ('ROI on SP — $2 win (%)',
         f"=SUM({s2_range('AA')})/({n}*2)*100",
         f"=SUM({s2_range('Z')})/({n}*2)*100"),
        ('Agreement Rate (%)',
         f"=COUNTIF({s2_range('O')},\"Y\")/COUNTA({s2_range('O')})*100",
         '—'),
        ('Disagreement Winner: CM right',
         f"=COUNTIFS({s2_range('O')},\"N\",{s2_range('U')},1)",
         '—'),
        ('Disagreement Winner: R5 right',
         f"=COUNTIFS({s2_range('O')},\"N\",{s2_range('T')},1)",
         '—'),
        ('Disagreement Winner: Tie/Other',
         f"=COUNTIFS({s2_range('O')},\"N\",{s2_range('T')},0,{s2_range('U')},0)",
         '—'),
        ('A-tier Hit Rate (%)',
         f"=COUNTIFS({s2_range('L')},\"A\",{s2_range('U')},1)/COUNTIF({s2_range('L')},\"A\")*100",
         '—'),
        ('HIGH-tier Hit Rate (%)',
         '—',
         f"=COUNTIFS({s2_range('I')},\"HIGH\",{s2_range('T')},1)/COUNTIF({s2_range('I')},\"HIGH\")*100"),
        (f'Race Count', f'={n}', f'={n}'),
    ]

    for ri, (label, cm_formula, r5_formula) in enumerate(metrics, 2):
        ws1.cell(ri, 1, label).font = BOLD
        ws1.cell(ri, 2, cm_formula)
        ws1.cell(ri, 3, r5_formula)

    ws1.freeze_panes = 'A2'

    # ================================================================
    # SHEET 3 — Breakdowns (SUMIF/COUNTIF formulas referencing Sheet 2)
    # ================================================================
    ws3 = wb.create_sheet("Breakdowns")

    def breakdown_block(ws, start_row, label, group_col_letter, group_values, block_label):
        """Write a breakdown block with SUMIF/COUNTIF formulas."""
        ws.cell(start_row, 1, block_label).font = Font(bold=True, size=12)
        start_row += 1
        headers = ['Group', 'Races', 'CM Wins', 'CM Win%', 'R5 Wins', 'R5 Win%',
                   'CM ROI(ML)%', 'R5 ROI(ML)%', 'CM ROI(SP)%', 'R5 ROI(SP)%', 'Agreement%']
        for ci, h in enumerate(headers, 1):
            hdr_cell(ws, start_row, ci, h, HDR_FILL_GREY)
        start_row += 1

        gcol = f"'Race by Race'!{group_col_letter}2:'{group_col_letter}{last_row}"  # for range ref

        for val in group_values:
            r = start_row
            ws.cell(r, 1, val)
            grange = f"'Race by Race'!{group_col_letter}$2:{group_col_letter}${last_row}"
            crange_u = f"'Race by Race'!U$2:U${last_row}"
            crange_t = f"'Race by Race'!T$2:T${last_row}"
            crange_y = f"'Race by Race'!Y$2:Y${last_row}"
            crange_x = f"'Race by Race'!X$2:X${last_row}"
            crange_aa= f"'Race by Race'!AA$2:AA${last_row}"
            crange_z = f"'Race by Race'!Z$2:Z${last_row}"
            crange_o = f"'Race by Race'!O$2:O${last_row}"

            criteria = f'"{val}"'
            ws.cell(r, 2, f"=COUNTIF({grange},{criteria})")
            ws.cell(r, 3, f"=COUNTIFS({grange},{criteria},{crange_u},1)")
            ws.cell(r, 4, f"=IF(COUNTIF({grange},{criteria})=0,\"\",COUNTIFS({grange},{criteria},{crange_u},1)/COUNTIF({grange},{criteria})*100)")
            ws.cell(r, 5, f"=COUNTIFS({grange},{criteria},{crange_t},1)")
            ws.cell(r, 6, f"=IF(COUNTIF({grange},{criteria})=0,\"\",COUNTIFS({grange},{criteria},{crange_t},1)/COUNTIF({grange},{criteria})*100)")
            ws.cell(r, 7, f"=IF(COUNTIF({grange},{criteria})=0,\"\",SUMIF({grange},{criteria},{crange_y})/(COUNTIF({grange},{criteria})*2)*100)")
            ws.cell(r, 8, f"=IF(COUNTIF({grange},{criteria})=0,\"\",SUMIF({grange},{criteria},{crange_x})/(COUNTIF({grange},{criteria})*2)*100)")
            ws.cell(r, 9, f"=IF(COUNTIF({grange},{criteria})=0,\"\",SUMIF({grange},{criteria},{crange_aa})/(COUNTIF({grange},{criteria})*2)*100)")
            ws.cell(r, 10,f"=IF(COUNTIF({grange},{criteria})=0,\"\",SUMIF({grange},{criteria},{crange_z})/(COUNTIF({grange},{criteria})*2)*100)")
            ws.cell(r, 11,f"=IF(COUNTIF({grange},{criteria})=0,\"\",COUNTIFS({grange},{criteria},{crange_o},\"Y\")/COUNTIF({grange},{criteria})*100)")
            start_row += 1

        return start_row + 1

    ws3.column_dimensions['A'].width = 16
    for ci in range(2, 12):
        ws3.column_dimensions[get_column_letter(ci)].width = 13

    row = 1
    # Unique values from data
    tracks   = sorted(set(r['track'] for r in data))
    surfaces = sorted(set(str(r['surface'] or '') for r in data if r['surface']))
    rtypes   = sorted(set(str(r['race_type'] or '') for r in data if r['race_type']))
    pscens   = sorted(set(str(r['pace_scenario'] or '') for r in data if r['pace_scenario']))

    row = breakdown_block(ws3, row, 'Track',          'B', tracks,   'By Track')
    row = breakdown_block(ws3, row, 'Surface',        'D', surfaces, 'By Surface')
    row = breakdown_block(ws3, row, 'Race Type',      'F', rtypes,   'By Race Type')
    row = breakdown_block(ws3, row, 'Pace Scenario',  'G', pscens,   'By Pace Scenario')

    # Field size: derive from counts in CM picks
    cm_con = sqlite3.connect(f"file:{os.path.abspath(CM_DB)}?mode=ro", uri=True)
    cm_cur = cm_con.cursor()
    cm_cur.execute("""
        SELECT track, race_date, race, COUNT(*) as field_size
        FROM picks GROUP BY track, race_date, race
    """)
    field_sizes = {}
    for rw in cm_cur.fetchall():
        field_sizes[(rw[0], rw[1], rw[2])] = rw[3]
    cm_con.close()

    # Write field size as extra column in Sheet 2 (col AB = 28)
    ws2.cell(1, 28, 'Field Size').font = BOLD
    for ri, row_d in enumerate(data, 2):
        fs = field_sizes.get((row_d['track'], row_d['date'], row_d['race']), '')
        ws2.cell(ri, 28, fs)

    # Field size breakdown using AB column
    ws3_row = row
    fs_groups = ['1-6', '7-9', '10+']
    ws3.cell(ws3_row, 1, 'By Field Size').font = Font(bold=True, size=12)
    ws3_row += 1
    fs_headers = ['Group', 'Races', 'CM Wins', 'CM Win%', 'R5 Wins', 'R5 Win%']
    for ci, h in enumerate(fs_headers, 1):
        hdr_cell(ws3, ws3_row, ci, h, HDR_FILL_GREY)
    ws3_row += 1

    for fs_label, fs_data in [
        ('1-6',  [r for r in data if field_sizes.get((r['track'], r['date'], r['race']), 0) <= 6]),
        ('7-9',  [r for r in data if 7 <= field_sizes.get((r['track'], r['date'], r['race']), 0) <= 9]),
        ('10+',  [r for r in data if field_sizes.get((r['track'], r['date'], r['race']), 0) >= 10]),
    ]:
        total = len(fs_data)
        cm_w = sum(r['cm_won'] for r in fs_data)
        r5_w = sum(r['r5_won'] for r in fs_data)
        ws3.cell(ws3_row, 1, fs_label)
        ws3.cell(ws3_row, 2, total)
        ws3.cell(ws3_row, 3, cm_w)
        ws3.cell(ws3_row, 4, round(cm_w/total*100, 1) if total else '')
        ws3.cell(ws3_row, 5, r5_w)
        ws3.cell(ws3_row, 6, round(r5_w/total*100, 1) if total else '')
        ws3_row += 1

    # ================================================================
    # SHEET 4 — CM Signals
    # ================================================================
    ws4 = wb.create_sheet("CM Signals")
    ws4.column_dimensions['A'].width = 26
    ws4.column_dimensions['B'].width = 10
    ws4.column_dimensions['C'].width = 10
    ws4.column_dimensions['D'].width = 12

    r = 1
    ws4.cell(r, 1, 'CM Signal Analysis').font = Font(bold=True, size=13)
    r += 2

    # A-tier
    ws4.cell(r, 1, 'A-Tier (rank=1 horses)').font = BOLD
    r += 1
    for lbl, val in [('Total', signals['a_tier']['total']),
                     ('Wins',  signals['a_tier']['wins']),
                     ('Win %', round(signals['a_tier']['rate'], 1))]:
        ws4.cell(r, 1, lbl)
        ws4.cell(r, 2, val)
        r += 1
    r += 1

    # Dominant flag
    ws4.cell(r, 1, 'Dominant Flag').font = BOLD
    r += 1
    for lbl, val in [('Total', signals['dominant']['total']),
                     ('Wins',  signals['dominant']['wins']),
                     ('Win %', round(signals['dominant']['rate'], 1))]:
        ws4.cell(r, 1, lbl)
        ws4.cell(r, 2, val)
        r += 1
    r += 1

    # Overlay Watch
    ws4.cell(r, 1, 'Overlay Watch').font = BOLD
    r += 1
    for lbl, val in [('Total', signals['overlay']['total']),
                     ('Wins',  signals['overlay']['wins']),
                     ('Win %', round(signals['overlay']['rate'], 1)),
                     ('ROI ML %', round(signals['overlay']['roi'], 1))]:
        ws4.cell(r, 1, lbl)
        ws4.cell(r, 2, val)
        r += 1
    r += 1

    # Consensus count vs win rate
    ws4.cell(r, 1, 'Consensus Count vs Win Rate (CM Rank=1)').font = BOLD
    r += 1
    for h, ci in [('Level', 1), ('Races', 2), ('Wins', 3), ('Win %', 4)]:
        hdr_cell(ws4, r, ci, h, HDR_FILL_GREY)
    r += 1
    for level in range(1, 9):
        s = signals['consensus'].get(level, {'total': 0, 'wins': 0, 'rate': 0})
        ws4.cell(r, 1, level)
        ws4.cell(r, 2, s['total'])
        ws4.cell(r, 3, s['wins'])
        ws4.cell(r, 4, round(s['rate'], 1))
        r += 1
    r += 1

    # Per-category underline hit rate
    ws4.cell(r, 1, 'Underline Hit Rate by Category (rank=1 underlined)').font = BOLD
    r += 1
    for h, ci in [('Category', 1), ('Total', 2), ('Wins', 3), ('Win %', 4)]:
        hdr_cell(ws4, r, ci, h, HDR_FILL_GREY)
    r += 1
    for cat, s in sorted(signals['underline'].items()):
        ws4.cell(r, 1, cat)
        ws4.cell(r, 2, s['total'])
        ws4.cell(r, 3, s['wins'])
        ws4.cell(r, 4, round(s['rate'], 1))
        r += 1

    # ================================================================
    # SHEET 5 — Disagreement Cases
    # ================================================================
    ws5 = wb.create_sheet("Disagreement Cases")
    disagreements = load_disagreements(data)

    s5_headers = ['Date', 'Track', 'Race', 'R5 Pick', 'R5 Score',
                  'CM Pick', 'CM Composite', 'Winner', 'Who Was Right', 'SP']
    for ci, h in enumerate(s5_headers, 1):
        hdr_cell(ws5, 1, ci, h, HDR_FILL_PURP)

    for ri, row_d in enumerate(disagreements, 2):
        r5_right = row_d['r5_won'] == 1
        cm_right = row_d['cm_won'] == 1
        who = 'R5' if r5_right else ('CM' if cm_right else 'Neither')
        ws5.cell(ri, 1, row_d['date'])
        ws5.cell(ri, 2, row_d['track'])
        ws5.cell(ri, 3, row_d['race'])
        ws5.cell(ri, 4, row_d['r5_pgm'])
        ws5.cell(ri, 5, row_d['r5_score'])
        ws5.cell(ri, 6, row_d['cm_pgm'])
        ws5.cell(ri, 7, row_d['cm_score'])
        ws5.cell(ri, 8, row_d['winner_name'])
        ws5.cell(ri, 9, who)
        ws5.cell(ri, 10, row_d['sp'])

    set_col_widths(ws5, [12, 6, 5, 8, 8, 8, 11, 24, 12, 8])
    ws5.freeze_panes = 'A2'

    # Save
    wb.save(output_path)
    return output_path


def run_recalc(xlsx_path: str) -> dict:
    result = subprocess.run(
        [sys.executable, RECALC, xlsx_path],
        capture_output=True, text=True
    )
    try:
        return json.loads(result.stdout)
    except Exception:
        return {"status": "error", "errors": [result.stdout or result.stderr], "formula_count": 0}


def generate_report() -> str:
    ts = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    n = len(load_sheet2_data())
    out = os.path.join(REPORTS_DIR, f"comparemodels_vs_r5_{n}races_{ts}.xlsx")

    print(f"Building comparison report → {out}")
    build_report(out)
    print("Report written.")

    print("Running recalc.py …")
    rc = run_recalc(out)
    print(json.dumps(rc, indent=2))
    if rc.get('status') != 'success':
        print("HALT: recalc.py returned errors.")
        sys.exit(1)
    print(f"recalc OK — {rc.get('formula_count', 0)} formulas checked, 0 errors.")
    return out


if __name__ == '__main__':
    generate_report()
